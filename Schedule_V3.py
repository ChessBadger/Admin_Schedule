import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import json
import openpyxl
import os
import subprocess

json_file_path = 'store_runs.json'

# Check if the file exists and then delete it
if os.path.exists(json_file_path):
    os.remove(json_file_path)

    # Define the current directory
directory = '.'

# List all files in the directory
files_in_directory = os.listdir(directory)

# Filter out Excel files
excel_files = [file for file in files_in_directory if file.endswith(
    '.xlsx') or file.endswith('.xls')]

# Delete each Excel file
for file in excel_files:
    os.remove(os.path.join(directory, file))


class StoreRun:
    def __init__(self, date, meet_time, start_time, store_note=None, employee_list=None):
        self.date = date
        self.meet_time = meet_time
        self.start_time = start_time
        self.inv_type = []
        self.store_name = []
        self.address = []
        self.link = []
        self.store_note = store_note
        self.employee_list = {}

    def add_employee(self, name, number, note, office):
        self.employee_list[name] = [number, note, office]

    def add_inv_type(self, inv_type):
        self.inv_type.append(inv_type)

    def add_store_name(self, store_name):
        self.store_name.append(store_name)

    def add_address(self, address):
        self.address.append(address)

    def add_link(self, link):
        self.link.append(link)

    def __str__(self):
        return f"date={self.date}, meet_time={self.meet_time}, start_time={self.start_time}, inv_type={self.inv_type}, store_name={self.store_name}, address={self.address}, link={self.link}, store_note={self.store_note}, employee_list={self.employee_list})"

    def to_dict(self):
        return {
            "date": self.date,
            "meet_time": self.meet_time,
            "start_time": self.start_time,
            "inv_type": self.inv_type,
            "store_name": self.store_name,
            "address": self.address,
            "link": self.link,
            "store_note": self.store_note,
            "employee_list": self.employee_list,
        }


def save_store_runs_to_json():
    # after creating and populating the store_run object
    json_file_path = 'store_runs.json'

    try:
        with open(json_file_path, 'r+') as file:
            # First we load existing data into a dict.
            file_data = json.load(file)
            # Join new_data with file_data inside emp_details
            file_data.append(store_run.to_dict())
            # Sets file's current position at offset.
            file.seek(0)
            # convert back to json.
            json.dump(file_data, file, indent=4)
    except FileNotFoundError:
        with open(json_file_path, 'w') as file:
            # If the file doesn't exist, create a new one and write the data.
            json.dump([store_run.to_dict()], file, indent=4)


script_dir = os.path.dirname(__file__)
config_path = os.path.join(script_dir, 'config.json')

# Load credentials from a JSON file
with open(config_path) as config_file:
    credentials = json.load(config_file)

# Set up the credentials for Google Sheets and Drive API
scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets'
]
creds = Credentials.from_service_account_info(credentials, scopes=scope)

# Initialize the clients for Sheets and Drive API
client = gspread.authorize(creds)
drive_service = build('drive', 'v3', credentials=creds)

# The name of the folder you want to search for
folder_name = 'Shared Test'

# Search for the folder by name to get its ID
query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}'"
folder_response = drive_service.files().list(q=query).execute()
folders = folder_response.get('files', [])

if folders:
    folder_id = folders[0]['id']  # Taking the first folder found

    # List all sheets in the specified folder using the folder ID
    query = f"mimeType='application/vnd.google-apps.spreadsheet' and '{folder_id}' in parents"
    file_response = drive_service.files().list(q=query).execute()
    files = file_response.get('files', [])
    files = sorted(files, key=lambda x: x['name'])
    process_columns = [2, 6, 10, 14, 18, 22, 26]

    for file in files:
        # Open the spreadsheet by ID using gspread
        gsheet = client.open_by_key(file['id'])
        # Print the name of the spreadsheet
        print(f"Processing sheet: {gsheet.title}")

        # Select the worksheet by title
        worksheet = gsheet.sheet1

        # Extract data from Google Sheets
        data = worksheet.get_all_values()

        # Create a new Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Insert data into the Excel sheet maintaining formatting
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_data in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index).value = cell_data

        workbook.save(gsheet.title + '.xlsx')

        if workbook:
            for process_col in process_columns:

                # Starting from row 8 in column B
                min_row = 8
                current_state = 'searching'  # Initial state before setting any values
                break_outer_loop = False  # Flag to control breaking out of the outer loop
                header_value = sheet.cell(row=1, column=process_col - 1).value

                # Iterate through the rows starting from min_row and process_col
                for row in sheet.iter_rows(min_row=min_row, max_row=163, min_col=process_col, max_col=process_col):
                    if break_outer_loop:
                        break
                    for cell in row:
                        value = cell.value
                        if value:
                            value = value.strip()
                        # Get the value of the cell to the left (number value)
                        number_value = sheet.cell(
                            row=cell.row, column=cell.column - 1).value
                        # Get the value of the cell to the right (note value)
                        note_value = sheet.cell(
                            row=cell.row, column=cell.column + 1).value
                        # Get the value of the cell two rows down (next store)
                        next_store = sheet.cell(
                            row=cell.row + 2, column=cell.column).value

                        # Handle empty days (weekends)
                        if cell.row == 8 and cell.value is None:
                            next_cell = sheet.cell(
                                row=cell.row + 1, column=cell.column).value
                            if next_cell is None:
                                next_cell = sheet.cell(
                                    row=cell.row + 1, column=cell.column).value
                                if next_cell is None:
                                    store_run = StoreRun(
                                        date=None, meet_time=None, start_time=None)
                                    store_run.date = header_value
                                    save_store_runs_to_json()
                                    break_outer_loop = True
                                    break
                        elif cell.row == 8 and cell.value == '':
                            next_cell = sheet.cell(
                                row=cell.row + 1, column=cell.column).value
                            if next_cell == '':
                                next_cell = sheet.cell(
                                    row=cell.row + 1, column=cell.column).value
                                if next_cell == '':
                                    store_run = StoreRun(
                                        date=None, meet_time=None, start_time=None)
                                    store_run.date = header_value
                                    save_store_runs_to_json()
                                    break_outer_loop = True
                                    break

                            # Check if the current state is 'searching' and the cell contains 'meet'
                        if value and current_state == 'searching' and 'meet' in value.lower():
                            store_run = StoreRun(
                                date=None, meet_time=None, start_time=None)
                            store_run.meet_time = value
                            current_state = 'found_meet'
                        elif value and current_state == 'searching' and 'leave time' in value.lower():
                            store_run = StoreRun(
                                date=None, meet_time=None, start_time=None)
                            store_run.meet_time = value
                            current_state = 'found_meet'
                        elif value and current_state == 'found_meet':
                            store_run.start_time = value
                            store_run.date = header_value
                            current_state = 'found_start'
                        # Check for anyone scheduled in the office
                        elif value and current_state == 'searching' and 'office' in value.lower() and 'leave' not in value.lower():
                            store_run = StoreRun(
                                date=None, meet_time=None, start_time=None)
                            store_run.date = header_value
                            store_run.add_store_name(value)
                            current_state = 'empty_value'
                        # If the current state is 'searching' or 'found_meet', capture the start time
                        elif value and current_state == 'searching':
                            store_run = StoreRun(
                                date=None, meet_time=None, start_time=None)
                            store_run.start_time = value
                            store_run.date = header_value
                            current_state = 'found_start'
                        # If the current state is 'found_start', capture the inventory type
                        elif current_state == 'found_start':
                            store_run.add_inv_type(value)
                            current_state = 'found_inv_type'
                        # If the current state is 'found_inv_type', capture the store name
                        elif current_state == 'found_inv_type':
                            store_run.add_store_name(value)
                            current_state = 'found_store_name'
                        # If the current state is 'found_store_name', capture the address
                        elif current_state == 'found_store_name':
                            store_run.add_address(value)
                            current_state = 'found_store_address'
                        # If the current state is 'found_store_address', capture the link
                        elif current_state == 'found_store_address':
                            store_run.add_link(value)
                            current_state = 'found_store_link'
                        # If the current state is 'found_store_link', handle 'TO FOLLOW' or note values
                        elif current_state == 'found_store_link':
                            if value == 'TO FOLLOW':
                                current_state = 'to_follow'
                            elif value and 'APPROX' in value:
                                current_state = 'to_follow'
                            elif value:
                                store_run.store_note = value
                                current_state = 'found_store_note'
                            else:
                                current_state = 'empty_value'
                        # If the current state is 'to_follow', capture another inventory type
                        elif current_state == 'to_follow':
                            store_run.add_inv_type(value)
                            current_state = 'found_inv_type'
                        # If the current state is 'found_store_note', set state to 'empty_value'
                        elif current_state == 'found_store_note':
                            current_state = 'empty_value'
                        # If the current state is 'empty_value', capture employee details
                        elif current_state == 'empty_value':
                            if 'Office' in store_run.store_name:
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                                current_state = 'searching'
                                next_cell = sheet.cell(
                                    row=cell.row + 1, column=cell.column).value
                                if next_cell:
                                    store_run.add_employee(
                                        next_cell, number_value, note_value, 'none')
                                    save_store_runs_to_json()
                                else:
                                    save_store_runs_to_json()
                            else:
                                current_state = 'found_employee'
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                        # If the current state is 'found_employee', continue capturing employees
                        elif current_state == "found_employee":
                            if value:
                                store_run.add_employee(
                                    value, number_value, note_value, 'none')
                            elif next_store:
                                current_state = 'searching'
                                save_store_runs_to_json()
                            # elif value is None or value == '':
                            #     if header_value == 'Mon, Jun 24':
                            #         print("kill me")
                            else:
                                save_store_runs_to_json()
                                break_outer_loop = True


else:
    print(f"No folder found with the name {folder_name}")


# Load the JSON data from the files
with open('formatted_users.json', 'r') as f:
    users_data = json.load(f)

with open('store_runs.json', 'r') as f:
    store_runs_data = json.load(f)

# Create a mapping of employee display names (in lowercase) to their office locations
employee_office_mapping = {user['displayName'].lower(
): user['office'] for user in users_data if 'displayName' in user and 'office' in user}

# Replace 'none' values in the store_runs employee_list with corresponding office values from the mapping
for run in store_runs_data:
    if 'employee_list' in run:
        for employee, details in run['employee_list'].items():
            if employee.lower() in employee_office_mapping:
                details[-1] = employee_office_mapping[employee.lower()]

# Save the updated store_runs data back to a new JSON file
updated_store_runs_path = 'store_runs.json'
with open(updated_store_runs_path, 'w') as f:
    json.dump(store_runs_data, f, indent=4)

print(f'Updated file saved to {updated_store_runs_path}')


# # Add all changes to the staging area
# subprocess.run(['git', 'add', '.'], check=True)

# # Commit changes with a message that includes the current date and time
# commit_message = subprocess.check_output(
#     ['date', '+%Y-%m-%d %H:%M:%S']).decode('utf-8').strip()
# subprocess.run(
#     ['git', 'commit', '-m', f"Automated commit. Updated: {commit_message}"], check=True)

# # Push changes to the 'main' branch of the 'origin' remote repository
# subprocess.run(['git', 'push', 'origin', 'main'], check=True)
